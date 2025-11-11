"""交叉对比门窗表和工程量清单"""
import os
from typing import Optional

import openpyxl
import re
from dataclasses import dataclass, field

from main import clean_str

DOOR_SHEET = r"./门窗表.xlsx"
BILL_SHEET = r"./工程量清单.xlsx"


@dataclass
class doorData:
    name: str
    facing: str = field(default="N/A")
    window: bool = field(default=False)
    num: int = field(default=0)
    from_path: Optional[str] = field(default=None)


def load_door_data() -> dict[str, doorData]:
    """加载门窗表数据"""
    door_data_: dict[str, doorData] = dict()

    print(f"正在处理门窗表: {DOOR_SHEET}")

    wb = openpyxl.load_workbook(DOOR_SHEET, data_only=True, read_only=True)

    # ws_list = [("1#2#办公楼地下部分门窗表", 8, 10), ("1#2#人防主要出入口门窗表", 5, 7)]
    # ws_list = [("1#办公楼地上部分门窗表", 17, 19)]
    ws_list = [("2#办公楼地上部分门窗表", 17, 19)]

    for sheet_name, col_index, facing_index in ws_list:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[2] is not None and row[col_index] is not None and int(row[col_index]) > 0
                    and "w" not in str(row[2]).lower() and "sh" not in str(row[2]).lower()
                    and "XCA" not in str(row[2]).upper()) and "#" not in str(row[2]):
                cleaned_key = clean_str(row[2])

                # 构造doorData对象
                door_data_[cleaned_key] = doorData(
                    name=cleaned_key, num=int(row[col_index]),
                    facing=row[facing_index] if row[facing_index] and "/" not in row[facing_index] else (
                        row[facing_index].split("/")[0].strip() if row[facing_index] else "N/A"),
                    window=("观察窗" in row[facing_index] if row[facing_index] else ""))

    # 检查door_data_中的是否存在重复
    duplicates = set([x for x in door_data_.keys() if list(door_data_.keys()).count(x) > 1])
    if duplicates:
        print(f"门窗表: {os.path.basename(DOOR_SHEET)}中存在重复项: {duplicates}")

    return door_data_


def load_bill_data(bill_sheet: str) -> dict[str, doorData]:
    """加载工程量清单数据"""
    print(f"正在处理工程量清单: {bill_sheet}")

    bill_data_: dict[str, doorData] = dict()

    wb = openpyxl.load_workbook(bill_sheet, data_only=True, read_only=True)
    ws = [sheet for sheet in wb.worksheets if "建筑工程" in sheet.title][0]

    key_list = []

    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[3] is not None and row[6] is not None and int(row[6]) > 0:
            cleaned_key = clean_bill_str(row[3], allowed_chinese=["防盗"])
            print(re.sub(r"\s+", " ", f"原始名称: {row[3]} -> 清洗后代号: {cleaned_key}"))
            key_list.append(cleaned_key)

            # 构造doorData对象
            bill_data_[cleaned_key] = doorData(
                name=cleaned_key, num=int(row[6]),
                facing="木纹" if "木制面" in row[3] else ("不锈钢" if "不锈钢" in row[3] else "粉末"),
                window=("观察窗" in row[3]),
                from_path=os.path.basename(bill_sheet) if bill_sheet else None
            )

    # 检查key_list中的是否存在重复
    duplicates = set([x for x in key_list if key_list.count(x) > 1])
    if duplicates:
        print(f"工程量清单: {os.path.basename(bill_sheet)}中存在重复项: {duplicates}")

    return bill_data_


def clean_bill_str(text: str, allowed_chinese: list[str]) -> str:
    """从给定字符串中提取清洗后的门型代号部分。"""
    if text is None:
        return ""
    text = str(text).replace("（", "(").replace("）", ")").replace("‘", "'").replace("’", "'").strip()
    # 定位并截取“名称：”之后的部分
    if "名称：" not in text:
        return ""
    rest = text.split("名称：", 1)[1]
    # 将换行符替换为空格，避免内容被截断
    rest = rest.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')

    output_chars = []
    inside_parentheses = False
    allowed_list = allowed_chinese if allowed_chinese is not None else []

    i = 0
    while i < len(rest):
        ch = rest[i]
        if inside_parentheses:
            # 处理括号内内容
            if ch == ')':  # 括号闭合
                inside_parentheses = False
                output_chars.append(ch)
                i += 1
                continue
            # 保留括号内的字母和数字
            if re.match(r'[0-9A-Za-z]', ch):
                output_chars.append(ch)
                i += 1
                continue
            # 括号内的中文字符：检查是否属于允许的关键词
            if '\u4e00' <= ch <= '\u9fff':
                matched = False
                for word in allowed_list:
                    # 如果从当前位置开始的子串匹配一个允许的关键词
                    if rest.startswith(word, i):
                        output_chars.append(word)
                        i += len(word)  # 跳过整个关键词
                        matched = True
                        break
                if matched:
                    continue  # 匹配到允许的词，已经处理完本轮
                else:
                    i += 1  # 非允许字符，跳过
                    continue
            # 其他字符（如空格或符号）在括号内则直接跳过
            i += 1
            continue
        else:
            # 处理括号外内容
            if ch == '(':
                inside_parentheses = True
                output_chars.append(ch)
                i += 1
                continue
            if '\u4e00' <= ch <= '\u9fff':
                # 括号外遇到中文，表示代号部分结束，停止提取
                break
            if ch.isspace():
                # 括号外的空白字符跳过（不计入代号，也不终止，以处理间隔）
                i += 1
                continue
            # 保留括号外的英文字母、数字、“.”、“-”、“_”等常见符号
            if re.match(r"[0-9A-Za-z.\-_']", ch):
                output_chars.append(ch)
                i += 1
                continue
            # 其他非预期字符视作代号结尾，退出循环
            break

    code = "".join(output_chars)
    # 若括号内容全被过滤掉，去除空的 "()"
    code = code.replace("()", "")
    return code


@dataclass
class cleanDictItem:
    # wb_path: str
    idx: int
    door_name: str
    sheet_name: str


@dataclass
class compareItem:
    name: str
    door_num: int = field(default=0)
    bill_num: int = field(default=0)
    door_facing: str = field(default="N/A")
    bill_facing: str = field(default="N/A")
    door_window: bool = field(default=False)
    bill_window: bool = field(default=False)


if __name__ == '__main__':
    door_data = load_door_data()

    bill_data: dict[str, doorData] = dict()

    for _bill_sheet in BILL_SHEET:
        data = load_bill_data(_bill_sheet)

        for key, value in data.items():
            print(f"项目: {key} 数量: {value.num}, 饰面: {value.facing}, 观察窗: {'有' if value.window else '无'}")

        bill_data.update(data)

    # 交叉对比两个dict
    print("-------------------开始数量交叉对比-------------------")
    all_keys = set(door_data.keys()).union(set(bill_data.keys()))
    compare_result: dict[str, compareItem] = dict()
    for key in all_keys:
        door_entry = door_data.get(key)
        bill_entry = bill_data.get(key)
        door_qty = door_entry.num if door_entry else 0
        bill_qty = bill_entry.num if bill_entry else 0
        if door_qty != bill_qty:
            compare_result[key] = compareItem(
                name=key,
                door_num=door_qty,
                bill_num=bill_qty,
                door_facing=door_entry.facing if door_entry else "N/A",
                bill_facing=bill_entry.facing if bill_entry else "N/A",
                door_window=door_entry.window if door_entry else False,
                bill_window=bill_entry.window if bill_entry else False
            )

        else:
            print(
                f"项目: {key} 数量一致: {door_qty}, 门窗表饰面: {door_entry.facing if door_entry else 'N/A'}, 观察窗: {'有' if door_entry and door_entry.window else '无'} | "
                f"工程量清单饰面: {bill_entry.facing if bill_entry else 'N/A'}, 观察窗: {'有' if bill_entry and bill_entry.window else '无'}")

    # 根据name排序compare_dict
    compare_result = dict(sorted(compare_result.items(), key=lambda x: x[0]))
    print("-------------------数量差异-------------------")
    for item in compare_result.values():
        if item.bill_num == 0:
            print(
                f"项目: {item.name} 在工程量清单中未找到对应项，门窗表数量: {item.door_num} (饰面: {item.door_facing}, 观察窗: {'有' if item.door_window else '无'})")
            # pass
        elif item.door_num == 0:
            print(
                f"项目: {item.name} 在门窗表中未找到对应项，工程量清单数量: {item.bill_num} (饰面: {item.bill_facing}, 观察窗: {'有' if item.bill_window else '无'})")
            # pass
        else:
            print(
                f"项目: {item.name} | 门窗表数量: {item.door_num} (饰面: {item.door_facing}, 观察窗: {'有' if item.door_window else '无'}) | "
                f"工程量清单数量: {item.bill_num} (饰面: {item.bill_facing}, 观察窗: {'有' if item.bill_window else '无'}) | "
                f"差异: {item.door_num - item.bill_num}")

    print("数量交叉对比完成")

    # total_itemized_dict: dict[str, cleanDictItem] = dict()
    #
    # for _bill_sheet in BILL_SHEET:
    #     itemized_dict: dict[str, cleanDictItem] = dict()
    #
    #     wb = openpyxl.load_workbook(_bill_sheet, read_only=True, data_only=True)
    #     ws_list = [clean_str(sheet.title).replace(" ", "") for sheet in wb.worksheets]
    #     for idx, sheet in [(i, sheet) for i, sheet in enumerate(wb.worksheets) if not bool(re.search(r'[\u4e00-\u9fa5]', sheet.title))]:
    #         cleaned_door_name = clean_str(sheet.cell(row=8, column=2).value).replace(" ", "")
    #         itemized_dict[cleaned_door_name] = cleanDictItem(idx=idx, door_name=cleaned_door_name, sheet_name=sheet.title)
    #
    #     print(f"工程量清单: {os.path.basename(_bill_sheet)}中分项有: {list(itemized_dict.keys())}")
    #
    #     total_itemized_dict.update(itemized_dict)
    #
    # # 对比total_itemized_dict.keys()和door_data.keys()
    # all_keys = set(total_itemized_dict.keys()).union(set(door_data.keys()))
    # for key in all_keys:
    #     itemized_entry = total_itemized_dict.get(key)
    #     door_entry = door_data.get(key)
    #     if door_entry and not itemized_entry:
    #         print(f"门窗表项目: {door_entry.name} 在工程量清单分项中未找到对应项")
