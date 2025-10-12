"""
used to create the drawing list with business requirements

输出目录结构:
    建筑重计量图纸目录
      ├─ 1.图纸确认
      ├─ 2.图纸审核证明
      └─ 建筑重计量图纸目录.xlsx
"""
import base64
import json
import os.path
import time
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.print_page_options import PrintOptions
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

from config import config
from dataclass import patternInfo
from main import requestToken, searchMail, MAIN_RE
from main_download_attachments import viewMailMetadata

XLSX_PATH: str = r"./图纸进度跟踪表.xlsx"
EXPORT_PATH: str = r"./建筑重计量图纸目录/建筑重计量图纸目录.xlsx"

# 邮件缓存路径
MAIL_CACHE_PATH = r"./cache/mail"
PROFILE_DIR = os.path.abspath("./cache/chrome_profile")


@dataclass
class DrawingItem:
    first_subject: str
    first_mail_id: int
    second_subject: str
    second_mail_id: int
    attachments: list[str] = field(default_factory=list)


def get_drawing_list() -> list[DrawingItem]:
    """get the drawing list from the xlsx file"""
    global MAIL_CACHE_PATH

    def get_row_data(search_params: patternInfo) -> DrawingItem:
        """get the row data from the search params"""
        search_response_list = searchMail(search_params=search_params)

        # init info_dict
        drawing_item = DrawingItem(
            first_subject=search_response_list[0].subject,
            first_mail_id=search_response_list[0].mailID,
            second_subject=search_response_list[1].subject.split(") ")[1] if len(search_response_list) > 1 else "",
            second_mail_id=search_response_list[1].mailID if len(search_response_list) > 1 else -1,
            attachments=[],
        )

        mail_response = viewMailMetadata(mail_id=drawing_item.first_mail_id)
        print(f"邮件: {mail_response.subject} ({mail_response.mail_id})")
        for _att in mail_response.attachments:
            drawing_item.attachments.append(os.path.splitext(_att.file_name)[0])
            print(f"  附件: {_att.file_name} ({_att.attachment_id})")
        print(drawing_item.first_subject)

        return drawing_item

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["自施范围(建筑装饰、门窗及室外工程)"]

    # init
    _info_list: list[DrawingItem] = []

    for row in ws.iter_rows(min_row=24):
        if row[1].value is not None:
            matched = MAIN_RE.match(row[1].value).groupdict()
            _data = get_row_data(search_params=patternInfo(
                unit=matched.get("unit"),
                discipline=matched.get("discipline"),
                drawing=matched.get("drawing"),
                step=matched.get("step"),
            ))
            _info_list.append(_data)
            # write to local json
            with open(rf"{MAIL_CACHE_PATH}/{_data.first_mail_id}.json", "w", encoding="utf-8") as _f:
                json.dump(_data.__dict__, _f, sort_keys=True, ensure_ascii=False, indent=4)
    wb.close()

    # 写图纸目录
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "建筑重计量图纸目录"
    ws.append(["序号", "图名", "图号"])  # 表头
    idx = 1
    for _item in _info_list:
        for att in _item.attachments:
            ws.append([idx, att, _item.first_subject])
            idx += 1
        # 合并图号列
        if len(_item.attachments) > 1:
            ws.merge_cells("C{}:C{}".format(idx - len(_item.attachments) + 1, idx))

    wb.save(EXPORT_PATH)
    wb.close()
    return _info_list


def get_driver() -> WebDriver:
    """get the selenium web driver"""
    global PROFILE_DIR
    chrome_options = Options()
    # 设置用户数据目录
    chrome_options.add_argument(f"--user-data-dir={PROFILE_DIR}")

    # 常用设置
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-gpu')

    # 禁用默认关闭
    # chrome_options.add_experimental_option("detach", True)

    # 隐藏特征
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument(
        '--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/109.0.5414.74 Safari/537.36')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.page_load_strategy = "normal"

    _driver = webdriver.Chrome(options=chrome_options, service=ChromeService(ChromeDriverManager().install()))
    # 隐藏特征
    _driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": """
            Object.defineProperty(navigator, 'webdriver', {
              get: () => undefined
            })
          """})
    # 抑制打印弹窗
    _driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {
            "source": """
              (function() {
                // 让 document.execCommand('print') 永远返回 false
                const origExec = Document.prototype.execCommand;
                Document.prototype.execCommand = function(cmd) {
                  if (cmd === 'print') { return false; }
                  return origExec.apply(this, arguments);
                };

                // 覆盖 window.print（包含所有同源 frame）
                const suppress = () => console.log('print() suppressed');
                Object.defineProperty(window, 'print', { value: suppress, writable: false });
              })();
            """
        },
    )

    return _driver


def get_mail_pdf(web_driver: WebDriver, mail_id: int) -> bytes:
    """get the mail pdf by mail id"""
    # 打开邮件页面-1
    web_driver.get(
        f"https://asia1.aconex.com/rsrc/20251003.0424/zh_CN_DOC/mail/view/index.html#/{config.project_id}/{mail_id}")
    print(f"正在打开邮件 {mail_id} 页面...")
    # 等待页面完全加载
    WebDriverWait(web_driver, timeout=10).until(
        EC.element_to_be_clickable((By.XPATH, "//a[@ng-click='toggleCollapsed()' and normalize-space(.)='消息']")))

    time.sleep(2)
    WebDriverWait(web_driver, timeout=3).until(
        EC.element_to_be_clickable((By.XPATH, "//button[normalize-space(.) = '打印']")))
    print_element = web_driver.find_element(By.XPATH, "//button[normalize-space(.) = '打印']")
    ActionChains(driver=web_driver).move_to_element(print_element).click(print_element).perform()

    time.sleep(2)
    WebDriverWait(web_driver, timeout=3).until(
        EC.element_to_be_clickable((By.XPATH, "//a[@data-automation-id='mailNavBar-printScreenModeNoThread']")))
    no_thread_element = web_driver.find_element(By.XPATH, "//a[@data-automation-id='mailNavBar-printScreenModeNoThread']")
    ActionChains(driver=web_driver).move_to_element(no_thread_element).click(no_thread_element).perform()

    # 打印页面
    print_options = PrintOptions()
    _pdf = web_driver.print_page(print_options)
    return base64.b64decode(_pdf)


if __name__ == '__main__':
    requestToken()

    # 构造输出结构
    confirm_path = Path(r"./建筑重计量图纸目录/1.图纸确认")
    confirm_path.mkdir(parents=True, exist_ok=True)
    verify_path = Path(r"./建筑重计量图纸目录/2.图纸审核证明")
    verify_path.mkdir(parents=True, exist_ok=True)

    # 构造邮件缓存目录
    mail_cache = Path(MAIL_CACHE_PATH)
    mail_cache.mkdir(parents=True, exist_ok=True)

    # info_list = get_drawing_list()
    # exit()

    # 读取本地json
    info_list: list[DrawingItem] = []
    for file in os.listdir(MAIL_CACHE_PATH):
        if file.endswith(".json"):
            with open(rf"{MAIL_CACHE_PATH}/{file}", "r", encoding="utf-8") as f:
                data = json.load(f)
                info_list.append(DrawingItem(**data))
    print(f"共读取 {len(info_list)} 条数据")

    # 获取driver
    driver = get_driver()
    driver.get("https://asia1.aconex.com/authentication/index.html")     # 登录页
    input("请在打开的浏览器中登录 Aconex 后，按回车继续...")

    for item in info_list:
        pdf_1 = get_mail_pdf(web_driver=driver, mail_id=item.first_mail_id)
        save_path = confirm_path / rf"{item.first_subject}.pdf"
        with open(save_path, "wb") as f:
            f.write(pdf_1)
        print(f"已保存邮件 {item.first_subject} ({item.first_mail_id}) 到 {save_path}")
        if item.second_mail_id == -1:
            print(f"邮件 {item.first_subject} 没有对应的审核证明，跳过")
            continue
        pdf_2 = get_mail_pdf(web_driver=driver, mail_id=item.second_mail_id)
        save_path = verify_path / rf"{item.second_subject}.pdf"
        with open(save_path, "wb") as f:
            f.write(pdf_2)
        print(f"已保存邮件 {item.second_subject} ({item.second_mail_id}) 到 {save_path}")
    driver.quit()
