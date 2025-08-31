import base64
import re
import threading
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone
from typing import Literal, Optional, List, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
import requests
from openpyxl.styles import PatternFill

from config import config
from dataclass import responseMailInfo, patternInfo, UserRef, WorkflowSearchResult, Workflow

ACCESS_TOKEN_LOCK = threading.Lock()

MAIN_RE = re.compile(
    r'^[ \t]*'                                        # ◇ 行首半角空白
    r'(?:.*?\((?P<wf>[A-Za-z]+-\d+)\)[ \t]*)?'        # ◇ 可选：(WF-001039)——前后可有任意文字
    r'.*?'                                            # ◇ 仍可再出现任意前缀（“通知：回复: 最终 ”等）
    r'SLDS-BCEG-'                                     # ▼ 固定文件名前缀
    r'(?P<unit>\d{3})-'                               # 单体
    r'SDS-'
    r'(?P<discipline>[A-Z]+)-'                        # 专业
    r'(?P<drawing>[A-Z0-9]+)'                         # 图纸号
    r'(?:_*(?P<ver>[A-Z]|\d+\+[A-Z]|\d+[A-Z]|\d+))?'  # 版本号（可选）
    r'(?:[ \t]+(?P<title>.+))?'                       # 图名（可选）
    r'[ \t]*$'                                        # ◇ 行尾半角空白
)

VER_RE = re.compile(r'^(?:(?P<num>\d+)(?:\+(?P<plus_letter>[A-Z])|(?P<letter>[A-Z])?)'
                    r'|(?P<pure_letter>[A-Z]))$')


def clean_str(s: str) -> str:
    """清理字符串"""
    return re.sub(r'\s+', ' ', s).replace('＿', '_').strip()


def sortMailsByVer(mails: list[responseMailInfo]) -> list[responseMailInfo]:
    """
    Sort the mails by version
    """

    def _ver_key(ver: str) -> tuple[int, int, int]:
        _m = VER_RE.match(ver) if ver else None
        if not _m:
            return 0, 4, 0  # 异常值，永远最后

        # ── 单字母 ──
        if _m.group('pure_letter'):  # L3
            return 0, 3, -ord(_m.group('pure_letter'))

        num_rank = -int(_m.group('num'))  # 数字降序
        if _m.group('plus_letter'):  # L0
            return num_rank, 0, -ord(_m.group('plus_letter'))
        if _m.group('letter'):  # L1
            return num_rank, 1, -ord(_m.group('letter'))
        return num_rank, 2, 0  # L2 纯数字

    def _mail_ver_key(mail):
        _m = MAIN_RE.match(mail.subject)
        ver = _m.group('ver') if (_m and _m.group('ver')) else ''
        return _ver_key(ver)

    return sorted(mails, key=_mail_ver_key)


def filter_mails(mails: List[responseMailInfo]) -> List[responseMailInfo]:
    """先按 mailID 去重，再按 (unit, discipline, drawing, ver) 聚合做优选"""

    # ---------- ① mailID 去重 ----------
    unique = {_m.mailID: _m for _m in mails}.values()

    # ---------- ② 内部工具函数 ----------
    def _parse_key(subj: str) -> Optional[Tuple[str, str, str, str]]:
        mo = MAIN_RE.search(subj)
        if not mo:
            return None
        return mo.group('unit'), mo.group('discipline'), mo.group('drawing'), mo.group('ver') or ''

    def _quality(subj: str) -> int:
        if '(WF-' in subj and not subj.startswith('最终'):
            return 2
        if '(WF-' in subj:
            return 1
        return 0

    def _better(prev: Optional[responseMailInfo], cand: responseMailInfo) -> responseMailInfo:
        """prev 允许为 None；返回质量更高 / 时间更新的邮件"""
        if prev is None:
            return cand
        k_prev = (_quality(prev.subject), prev.SentDate)
        k_cand = (_quality(cand.subject), cand.SentDate)
        return cand if k_cand > k_prev else prev

    # ---------- ③ 聚合并优选 ----------
    best: dict[Tuple[str, str, str, str], responseMailInfo] = {}
    for _m in unique:
        key = _parse_key(_m.subject)
        if key is None:  # subject 不满足规则可视需求忽略
            continue
        best[key] = _better(best.get(key), _m)  # 用 .get()，首轮 prev=None

    return list(best.values())


def responseClean(mail_response: list[responseMailInfo], search_params: patternInfo) -> list[responseMailInfo]:
    """
    Clean the mail response by filtering based title
    """
    subject = f"SLDS-BCEG-{search_params.unit}-SDS-{search_params.discipline}-{search_params.drawing}"
    # return sortMailsByVer([mail for mail in mail_response if mail.subject.startswith(subject)])
    return sortMailsByVer([mail for mail in mail_response if subject in mail.subject])


def requestToken() -> dict:
    """
    Request an OAuth2 token using client credentials grant type

    https://help.aconex.com/zh/apis/implement-smart-construction-platform-oauth/#Implement-OAuth-in-a-User-Bound-Integration
    """

    def basicAuthEncode(client_id: str, client_secret: str) -> str:
        """
        Encode client_id and client_secret in base64 for Basic Auth

        https://en.wikipedia.org/wiki/Basic_access_authentication
        """
        auth_str = f"{client_id}:{client_secret}"
        b64_encoded = base64.b64encode(auth_str.encode()).decode()
        return b64_encoded

    response = requests.post(url=f"{config.lobby_url}/auth/token",
                             headers={"Content-Type": "application/x-www-form-urlencoded",
                                      "Authorization": f"Basic {basicAuthEncode(config.client_id, config.client_secret)}"},
                             data={"grant_type": "client_credentials", "user_id": config.aconex_user_id,
                                   "user_site": config.aconex_instance_url},
                             proxies={"http": "127.0.0.1:52538", "https": "127.0.0.1:52538"})

    response.raise_for_status()

    config.access_token = response.json().get("access_token")
    config.access_token_expires = datetime.now() + timedelta(seconds=response.json().get("expires_in") - 300)
    return response.json()


def searchMail(search_params: patternInfo = patternInfo(unit="000", discipline="A", drawing="A001"),
               mail_box: Literal["INBOX", "SENTBOX", "ALL"] = "ALL") -> list[responseMailInfo]:
    """
    Search mails by subject

    https://help.aconex.com/zh/apis/mail-api-developer-guide/
    """

    def searchQueryCreator() -> str:
        """
        根据深化图编号规则生成 search_query 字符串
        unit     单体号, 3 位 (e.g. '001')
        discipline     专业代码, 大写 (e.g. 'HV')
        drawing  图号, 3 位 (e.g. '001')
        ver      版本号, 形如 '_0', '_A'；默认为 '*' 通配全部版本
        """
        # Lucene 表达式：拆分词后用 AND 组合 + 通配符
        tokens = f"{search_params.unit} {search_params.discipline} {search_params.drawing}{search_params.ver}".split()
        subject_cond = " AND ".join(f"{tok}" for tok in tokens)
        # query = rf"subject:({subject_cond}) AND corrtypeid:23"
        query = rf"subject:({subject_cond})"
        return query

    def responseMailInfoPostprocess(xml_text: bytes) -> list[responseMailInfo]:
        """
        处理返回的邮件数据，转换为 responseMailInfo 列表
        """
        root = ET.fromstring(xml_text.decode("utf-8"))
        export_list = []
        for _mail in root.find('SearchResults').iter('Mail'):
            # print(ET.tostring(m, encoding='utf-8').decode('utf-8'))
            export_list.append(responseMailInfo(mailID=int(_mail.attrib['MailId']), MailNo=_mail.findtext('MailNo'),
                                                SentDate=datetime.fromisoformat(_mail.findtext('SentDate').rstrip('Z')),
                                                subject=_mail.findtext('Subject'),
                                                AllAttachmentCount=int(_mail.findtext('AllAttachmentCount')), ))
        return export_list

    # 检查输入变量
    print(f"Search params: {search_params.__dict__}, mail box: {mail_box}")

    # 检查当前时间是否超过了config.access_token_expires，如果超过则调用requestToken()刷新token
    if not config.access_token_expires or not config.access_token or datetime.now() >= config.access_token_expires:
        with ACCESS_TOKEN_LOCK:
            print("Access token expired, refreshing...")
            requestToken()

    mail = []

    # SentBox
    if mail_box == "SENTBOX" or mail_box == "ALL":
        response = requests.get(url=f"{config.resource_url}/api/projects/{config.project_id}/mail",
                                headers={"Authorization": f"Bearer {config.access_token}"},
                                params={"mail_box": "SENTBOX", "search_query": searchQueryCreator(),
                                        "return_fields": "docno,subject,sentdate,allAttachmentCount,totalAttachmentsSize",
                                        "sort_field": "sentdate", "sort_direction": "DESC"},
                                proxies={"http": "127.0.0.1:52538", "https": "127.0.0.1:52538"})

        response.raise_for_status()

        mail += responseMailInfoPostprocess(response.content)

    # InBox
    if mail_box == "INBOX" or mail_box == "ALL":
        response = requests.get(url=f"{config.resource_url}/api/projects/{config.project_id}/mail",
                                headers={"Authorization": f"Bearer {config.access_token}"},
                                params={"mail_box": "INBOX", "search_query": searchQueryCreator(),
                                        "return_fields": "docno,subject,sentdate,allAttachmentCount,totalAttachmentsSize",
                                        "sort_field": "sentdate", "sort_direction": "DESC"},
                                proxies={"http": "127.0.0.1:52538", "https": "127.0.0.1:52538"})

        response.raise_for_status()
        mail += responseMailInfoPostprocess(response.content)

    # 使用 filter_mails 去重和择优
    mail = filter_mails(mail)

    # mail 排序
    mail = responseClean(mail_response=mail, search_params=search_params)

    return mail


def parseWorkflowSearch(xml_text: bytes) -> WorkflowSearchResult:
    def _parseDatetime(dt: str) -> Optional[datetime]:
        """
        把 RFC-3339 / ISO-8601 字符串转为 datetime，并转换到 UTC+8。
        - 原始 API 字段形如 '2025-08-29T08:38:39.839Z'（Z 表示 UTC）
        - 返回值例如 2025-08-29 16:38:39.839+08:00
        """

        # 将 'Z' 替换为 '+00:00'，构造成可被 fromisoformat 解析的字符串
        utc_dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        # 转换到东八区
        return utc_dt.astimezone(timezone(timedelta(hours=8)))

    def _parseUser(elem: ET.Element) -> UserRef:
        """解析 <Assignee> / <Initiator> / <Reviewer>"""
        return UserRef(organization_id=int(elem.findtext("OrganizationId")),
                       organization_name=elem.findtext("OrganizationName").strip(), name=elem.findtext("Name").strip(),
                       user_id=int(elem.findtext("UserId")), )

    root = ET.fromstring(xml_text.decode("utf-8"))

    # ---- 顶层分页元数据 ----
    meta = WorkflowSearchResult(current_page=int(root.attrib["CurrentPage"]), page_size=int(root.attrib["PageSize"]),
                                total_pages=int(root.attrib["TotalPages"]),
                                total_results=int(root.attrib["TotalResults"]),
                                total_results_on_page=int(root.attrib["TotalResultsOnPage"]), workflows=[], )

    # ---- 循环解析每条 <Workflow> ----
    for wf_elem in root.find("SearchResults").findall("Workflow"):
        # ─ Assignees (0-N) ─
        assignees = [_parseUser(a) for a in wf_elem.find("Assignees").findall("Assignee")]

        # ─ Initiator / Reviewer (可能缺省) ─
        initiator = _parseUser(wf_elem.find("Initiator"))
        reviewer_elem = wf_elem.find("Reviewer")
        reviewer = _parseUser(reviewer_elem) if reviewer_elem is not None else None

        wf = Workflow(workflow_id=int(wf_elem.attrib["WorkflowId"]), step_name=wf_elem.findtext("StepName").strip(),
                      step_outcome=wf_elem.findtext("StepOutcome").strip(),
                      step_status=wf_elem.findtext("StepStatus").strip(),

                      date_in=_parseDatetime(wf_elem.findtext("DateIn")) if wf_elem.findtext("DateIn") else None,
                      date_completed=_parseDatetime(wf_elem.findtext("DateCompleted")) if wf_elem.findtext(
                          "DateCompleted") else None,
                      date_due=_parseDatetime(wf_elem.findtext("DateDue")) if wf_elem.findtext("DateDue") else None,
                      days_late=int(wf_elem.findtext("DaysLate")), duration=float(wf_elem.findtext("Duration")),

                      document_number=wf_elem.findtext("DocumentNumber").strip(),
                      document_revision=wf_elem.findtext("DocumentRevision").strip(),
                      document_title=wf_elem.findtext("DocumentTitle").strip(),
                      document_version=int(wf_elem.findtext("DocumentVersion")),
                      file_name=wf_elem.findtext("FileName").strip(),
                      file_size=int(wf_elem.findtext("FileSize")),

                      initiator=initiator, reviewer=reviewer, assignees=assignees, )
        meta.workflows.append(wf)
    return meta


def searchWorkflow(workflow_num: str) -> WorkflowSearchResult:
    # 检查当前时间是否超过了config.access_token_expires，如果超过则调用requestToken()刷新token
    if not config.access_token_expires or not config.access_token or datetime.now() >= config.access_token_expires:
        with ACCESS_TOKEN_LOCK:
            print("Access token expired, refreshing...")
            requestToken()

    response = requests.get(url=f"{config.resource_url}/api/projects/{config.project_id}/workflows/search",
                            headers={"Authorization": f"Bearer {config.access_token}",
                                     "Accept": "application/vnd.aconex.workflow.v1+xml", },
                            params={"workflow_number": {workflow_num}, },
                            proxies={"http": "127.0.0.1:52538", "https": "127.0.0.1:52538"})

    response.raise_for_status()
    return parseWorkflowSearch(response.content)


def process_row(args):
    """
    单线程处理一行数据。为了避免直接操作 Workbook，
    仅返回『需要写回行的列号和值』的列表。
    """
    row, m = args  # m 是 MAIN_RE.match 的结果
    matched_data = m.groupdict()

    # ---- 1. 请求邮件 ----
    mails = searchMail(
        search_params=patternInfo(
            unit=matched_data['unit'],
            discipline=matched_data['discipline'],
            drawing=matched_data['drawing'],
        ),
        mail_box="ALL"
    )

    if not mails:
        return row, []  # 无数据，主线程可打印日志

    newest_mail = mails[0]
    newest_match = MAIN_RE.match(newest_mail.subject).groupdict()

    _writes = [
        (4, newest_match.get('ver', '')),  # ver
        (5, newest_mail.SentDate.date().isoformat()),  # sentDate
    ]

    if newest_match.get('wf') and not newest_match['ver'].isdigit():
        wf_num = newest_match['wf']
        wf_data = searchWorkflow(wf_num)
        _writes.append((7, wf_num))

        base_col = 8
        for wf in wf_data.workflows:
            if wf.step_name == "最终" and wf.step_outcome != "正等待处理":
                _writes.append((6, f"code {wf.step_outcome.split('-')[0]}"))
            if wf.step_outcome == "正等待处理":
                _writes.extend([
                    (base_col, wf.assignees[0].organization_name.split(" ")[0]),
                    (base_col + 1, wf.assignees[0].name.split(" ")[-1]),
                    (base_col + 2, wf.step_status),
                ])
                base_col += 3

    if newest_match.get('ver', '').isdigit():
        _writes.append(('fill_green', None))  # 自定义标记，主线程再做 fill

    return row, _writes


# ---------------- ② main 修改 ----------------
if __name__ == '__main__':
    requestToken()
    wb = openpyxl.load_workbook('test1.xlsx')

    green_fill = PatternFill("solid", fgColor="92D050")

    for sheet in wb.worksheets:
        rows = [
            (row, MAIN_RE.match(clean_str(row[1].value)))
            for row in sheet.iter_rows(min_row=2, max_col=50)
            if row[1].value and MAIN_RE.match(clean_str(row[1].value))
        ]

        with ThreadPoolExecutor(max_workers=20) as executor:
            futures = {executor.submit(process_row, item): item[0] for item in rows}

            for fut in as_completed(futures):
                row, writes = fut.result()  # 写操作放回主线程
                for col_idx, val in writes:
                    if col_idx == 'fill_green':
                        for i in range(8):  # 0-7 列涂色
                            row[i].fill = green_fill
                    else:
                        row[col_idx].value = val

    wb.save('test1_out.xlsx')
    wb.close()
