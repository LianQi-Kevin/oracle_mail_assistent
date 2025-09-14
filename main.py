import base64
import os.path
import re
import threading
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone
from typing import Literal, Optional, List, Tuple
from concurrent.futures import ThreadPoolExecutor, wait, ALL_COMPLETED

import openpyxl
import requests
from openpyxl.cell import MergedCell, Cell
# from openpyxl.cell import Cell
from openpyxl.styles import PatternFill, Border, Side
# from openpyxl.styles import Font

from config import config
from dataclass import responseMailInfo, patternInfo, UserRef, WorkflowSearchResult, Workflow, searchResult

# COLOR
BLUE: str = "00B0F0"
YELLOW: str = "FFFF00"
GREEN: str = "92D050"

# LOCK
ACCESS_TOKEN_LOCK = threading.Lock()
CELL_WRITE_LOCK = threading.Lock()

# PATH
XLSX_PATH = r"./图纸进度跟踪表.xlsx"
# EXPORT_PATH = rf"./{os.path.splitext(os.path.basename(XLSX_PATH))[0]}_out.xlsx"
EXPORT_PATH = XLSX_PATH

# GLOBAL VARS
REQUEST_DATA: dict[str, searchResult] = dict()
MAX_COL_USED = 0

MAIN_RE = re.compile(
    r'^[ \t]*'                                        # 行首半角空白
    r'(?:.*?\((?P<wf>[A-Za-z]+-\d+)\)[ \t]*)?'        # 可选：(WF-001039)——前后可有任意文字
    r'.*?'                                            # 仍可再出现任意前缀（“通知：回复: 最终 ”等）
    r'SLDS-BCEG-'                                     # 固定文件名前缀
    r'(?P<unit>\d{3})-'                               # 单体
    r'SDS-'
    r'(?P<discipline>[A-Z]+)-'                        # 专业
    r'(?P<drawing>[A-Z0-9]+)'                         # 图纸号
    r'(?:_*(?P<ver>[A-Z]|\d+\+[A-Z]|\d+[A-Z]|\d+))?'  # 版本号（可选）
    r'(?:[ \t]+(?P<title>.+))?'                       # 图名（可选）
    r'[ \t]*$'                                        # 行尾半角空白
)

VER_RE = re.compile(r'^(?:(?P<num>\d+)(?:\+(?P<plus_letter>[A-Z])|(?P<letter>[A-Z])?)'
                    r'|(?P<pure_letter>[A-Z]))$')


def clean_str(s: str) -> str:
    """清理字符串"""
    return re.sub(r'\s+', ' ', s).replace('＿', '_').strip()


def sortMailsByVer(mails: list[responseMailInfo]) -> list[responseMailInfo]:
    """
    Sort the mails by version.

    新优先级（同数字情况下）：
        1) 数字+字母   例：12A
        2) 纯数字       例：12
        3) 数字+‘+字母’ 例：12+A
        4) 纯字母       例：A
        5) 异常 / 无版本
    """

    def _ver_key(ver: str) -> tuple[int, int, int]:
        _m = VER_RE.match(ver) if ver else None
        if not _m:
            return 0, 4, 0  # 异常值，永远最后

        # ── 分类后再组合排序键 ──
        # 第一位：数字降序（取负）
        # 第二位：档位码（越小优先）
        # 第三位：字母逆序（Z→A），无字母时置 0
        num_rank = -int(_m.group('num')) if _m.group('num') else 0

        if _m.group('letter'):                # 数字+字母 —— 档位 0（最高）
            return num_rank, 0, -ord(_m.group('letter'))
        if _m.group('num') and not any((_m.group('plus_letter'),
                                        _m.group('letter'),
                                        _m.group('pure_letter'))):  # 纯数字 —— 档位 1
            return num_rank, 1, 0
        if _m.group('plus_letter'):           # 数字+‘+字母’ —— 档位 2
            return num_rank, 2, -ord(_m.group('plus_letter'))
        if _m.group('pure_letter'):           # 纯字母 —— 档位 3
            return 0, 3, -ord(_m.group('pure_letter'))

        return 0, 4, 0        # 理论兜底，不会触发

    def _mail_ver_key(mail: responseMailInfo):
        _m = MAIN_RE.match(mail.subject)
        ver = _m.group('ver') if (_m and _m.group('ver')) else ''
        return _ver_key(ver)

    return sorted(mails, key=_mail_ver_key)


def filter_mails(mails: List[responseMailInfo]) -> List[responseMailInfo]:
    """先按 mailID 去重，再按 (unit, discipline, drawing, ver) 聚合做优选"""

    # ---------- ① mailID 去重 ----------
    unique = {_m.mailID: _m for _m in mails}.values()

    # ---------- ② 内部工具函数 ----------
    def _parse_key(subj: str) -> Optional[Tuple[str, str, str, str]]:
        mo = MAIN_RE.search(subj)
        if not mo:
            return None
        return mo.group('unit'), mo.group('discipline'), mo.group('drawing'), mo.group('ver') or ''

    def _quality(subj: str) -> int:
        if '(WF-' in subj and not subj.startswith('最终'):
            return 2
        if '(WF-' in subj:
            return 1
        return 0

    def _better(prev: Optional[responseMailInfo], cand: responseMailInfo) -> responseMailInfo:
        """prev 允许为 None；返回质量更高 / 时间更新的邮件"""
        if prev is None:
            return cand
        k_prev = (_quality(prev.subject), prev.SentDate)
        k_cand = (_quality(cand.subject), cand.SentDate)
        return cand if k_cand > k_prev else prev

    # ---------- ③ 聚合并优选 ----------
    best: dict[Tuple[str, str, str, str], responseMailInfo] = {}
    for _m in unique:
        key = _parse_key(_m.subject)
        if key is None:  # subject 不满足规则可视需求忽略
            continue
        best[key] = _better(best.get(key), _m)  # 用 .get()，首轮 prev=None

    return list(best.values())


def responseClean(mail_response: list[responseMailInfo], search_params: patternInfo) -> list[responseMailInfo]:
    """
    Clean the mail response by filtering based title
    """
    subject = f"SLDS-BCEG-{search_params.unit}-SDS-{search_params.discipline}-{search_params.drawing}"
    # return sortMailsByVer([mail for mail in mail_response if mail.subject.startswith(subject)])
    return sortMailsByVer([mail for mail in mail_response if subject in mail.subject])


def requestToken() -> dict:
    """
    Request an OAuth2 token using client credentials grant type

    https://help.aconex.com/zh/apis/implement-smart-construction-platform-oauth/#Implement-OAuth-in-a-User-Bound-Integration
    """

    def basicAuthEncode(client_id: str, client_secret: str) -> str:
        """
        Encode client_id and client_secret in base64 for Basic Auth

        https://en.wikipedia.org/wiki/Basic_access_authentication
        """
        auth_str = f"{client_id}:{client_secret}"
        b64_encoded = base64.b64encode(auth_str.encode()).decode()
        return b64_encoded

    response = requests.post(url=f"{config.lobby_url}/auth/token",
                             headers={"Content-Type": "application/x-www-form-urlencoded",
                                      "Authorization": f"Basic {basicAuthEncode(config.client_id, config.client_secret)}"},
                             data={"grant_type": "client_credentials", "user_id": config.aconex_user_id,
                                   "user_site": config.aconex_instance_url},
                             proxies={"http": "127.0.0.1:52538", "https": "127.0.0.1:52538"})

    response.raise_for_status()

    config.access_token = response.json().get("access_token")
    config.access_token_expires = datetime.now() + timedelta(seconds=response.json().get("expires_in") - 300)
    return response.json()


def searchMail(search_params: patternInfo = patternInfo(unit="000", discipline="A", drawing="A001"),
               mail_box: Literal["INBOX", "SENTBOX", "ALL"] = "ALL") -> list[responseMailInfo]:
    """
    Search mails by subject

    https://help.aconex.com/zh/apis/mail-api-developer-guide/
    """
    global ACCESS_TOKEN_LOCK

    def searchQueryCreator() -> str:
        """
        根据深化图编号规则生成 search_query 字符串
        unit     单体号, 3 位 (e.g. '001')
        discipline     专业代码, 大写 (e.g. 'HV')
        drawing  图号, 3 位 (e.g. '001')
        ver      版本号, 形如 '_0', '_A'；默认为 '*' 通配全部版本
        """
        # Lucene 表达式：拆分词后用 AND 组合 + 通配符
        tokens = f"{search_params.unit} {search_params.discipline} {search_params.drawing}{search_params.ver}".split()
        subject_cond = " AND ".join(f"{tok}" for tok in tokens)
        # query = rf"subject:({subject_cond}) AND corrtypeid:23"
        query = rf"subject:({subject_cond})"
        return query

    def responseMailInfoPostprocess(xml_text: bytes) -> list[responseMailInfo]:
        """
        处理返回的邮件数据，转换为 responseMailInfo 列表
        """
        root = ET.fromstring(xml_text.decode("utf-8"))
        export_list = []
        for _mail in root.find('SearchResults').iter('Mail'):
            # print(ET.tostring(m, encoding='utf-8').decode('utf-8'))
            export_list.append(responseMailInfo(mailID=int(_mail.attrib['MailId']), MailNo=_mail.findtext('MailNo'),
                                                SentDate=datetime.fromisoformat(_mail.findtext('SentDate').rstrip('Z')),
                                                subject=_mail.findtext('Subject'),
                                                AllAttachmentCount=int(_mail.findtext('AllAttachmentCount')), ))
        return export_list

    # 检查输入变量
    print(f"Search params: {search_params.__dict__}, mail box: {mail_box}")

    # 检查当前时间是否超过了config.access_token_expires，如果超过则调用requestToken()刷新token
    if not config.access_token_expires or not config.access_token or datetime.now() >= config.access_token_expires:
        with ACCESS_TOKEN_LOCK:
            print("Access token expired, refreshing...")
            requestToken()

    mail = []

    # SentBox
    if mail_box == "SENTBOX" or mail_box == "ALL":
        response = requests.get(url=f"{config.resource_url}/api/projects/{config.project_id}/mail",
                                headers={"Authorization": f"Bearer {config.access_token}"},
                                params={"mail_box": "SENTBOX", "search_query": searchQueryCreator(),
                                        "return_fields": "docno,subject,sentdate,allAttachmentCount,totalAttachmentsSize",
                                        "sort_field": "sentdate", "sort_direction": "DESC"},
                                proxies={"http": "127.0.0.1:52538", "https": "127.0.0.1:52538"})

        response.raise_for_status()

        mail += responseMailInfoPostprocess(response.content)

    # InBox
    if mail_box == "INBOX" or mail_box == "ALL":
        response = requests.get(url=f"{config.resource_url}/api/projects/{config.project_id}/mail",
                                headers={"Authorization": f"Bearer {config.access_token}"},
                                params={"mail_box": "INBOX", "search_query": searchQueryCreator(),
                                        "return_fields": "docno,subject,sentdate,allAttachmentCount,totalAttachmentsSize",
                                        "sort_field": "sentdate", "sort_direction": "DESC"},
                                proxies={"http": "127.0.0.1:52538", "https": "127.0.0.1:52538"})

        response.raise_for_status()
        mail += responseMailInfoPostprocess(response.content)

    # 使用 filter_mails 去重和择优
    mail = filter_mails(mail)

    # mail 排序
    mail = responseClean(mail_response=mail, search_params=search_params)

    return mail


def parseWorkflowSearch(xml_text: bytes) -> WorkflowSearchResult:
    def _parseDatetime(dt: str) -> Optional[datetime]:
        """
        把 RFC-3339 / ISO-8601 字符串转为 datetime，并转换到 UTC+8。
        - 原始 API 字段形如 '2025-08-29T08:38:39.839Z'（Z 表示 UTC）
        - 返回值例如 2025-08-29 16:38:39.839+08:00
        """

        # 将 'Z' 替换为 '+00:00'，构造成可被 fromisoformat 解析的字符串
        utc_dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        # 转换到东八区
        return utc_dt.astimezone(timezone(timedelta(hours=8)))

    def _parseUser(elem: ET.Element) -> UserRef:
        """解析 <Assignee> / <Initiator> / <Reviewer>"""
        return UserRef(organization_id=int(elem.findtext("OrganizationId")),
                       organization_name=elem.findtext("OrganizationName").strip(), name=elem.findtext("Name").strip(),
                       user_id=int(elem.findtext("UserId")), )

    root = ET.fromstring(xml_text.decode("utf-8"))

    # ---- 顶层分页元数据 ----
    meta = WorkflowSearchResult(current_page=int(root.attrib["CurrentPage"]), page_size=int(root.attrib["PageSize"]),
                                total_pages=int(root.attrib["TotalPages"]),
                                total_results=int(root.attrib["TotalResults"]),
                                total_results_on_page=int(root.attrib["TotalResultsOnPage"]), workflows=[], )

    # ---- 循环解析每条 <Workflow> ----
    for wf_elem in root.find("SearchResults").findall("Workflow"):
        # ─ Assignees (0-N) ─
        assignees = [_parseUser(a) for a in wf_elem.find("Assignees").findall("Assignee")]

        # ─ Initiator / Reviewer (可能缺省) ─
        initiator = _parseUser(wf_elem.find("Initiator"))
        reviewer_elem = wf_elem.find("Reviewer")
        reviewer = _parseUser(reviewer_elem) if reviewer_elem is not None else None

        wf = Workflow(workflow_id=int(wf_elem.attrib["WorkflowId"]), step_name=wf_elem.findtext("StepName").strip(),
                      step_outcome=wf_elem.findtext("StepOutcome").strip(),
                      step_status=wf_elem.findtext("StepStatus").strip(),

                      date_in=_parseDatetime(wf_elem.findtext("DateIn")) if wf_elem.findtext("DateIn") else None,
                      date_completed=_parseDatetime(wf_elem.findtext("DateCompleted")) if wf_elem.findtext(
                          "DateCompleted") else None,
                      date_due=_parseDatetime(wf_elem.findtext("DateDue")) if wf_elem.findtext("DateDue") else None,
                      days_late=int(wf_elem.findtext("DaysLate")), duration=float(wf_elem.findtext("Duration")),

                      document_number=wf_elem.findtext("DocumentNumber").strip(),
                      document_revision=wf_elem.findtext("DocumentRevision").strip(),
                      document_title=wf_elem.findtext("DocumentTitle").strip(),
                      document_version=int(wf_elem.findtext("DocumentVersion")),
                      file_name=wf_elem.findtext("FileName").strip(),
                      file_size=int(wf_elem.findtext("FileSize")),

                      initiator=initiator, reviewer=reviewer, assignees=assignees, )
        meta.workflows.append(wf)
    return meta


def searchWorkflow(workflow_num: str) -> WorkflowSearchResult:
    # 检查当前时间是否超过了config.access_token_expires，如果超过则调用requestToken()刷新token
    if not config.access_token_expires or not config.access_token or datetime.now() >= config.access_token_expires:
        with ACCESS_TOKEN_LOCK:
            print("Access token expired, refreshing...")
            requestToken()

    response = requests.get(url=f"{config.resource_url}/api/projects/{config.project_id}/workflows/search",
                            headers={"Authorization": f"Bearer {config.access_token}",
                                     "Accept": "application/vnd.aconex.workflow.v1+xml", },
                            params={"workflow_number": {workflow_num}, },
                            proxies={"http": "127.0.0.1:52538", "https": "127.0.0.1:52538"})

    response.raise_for_status()
    return parseWorkflowSearch(response.content)


# def cellWriter(cell: Cell, value: Optional[str]):
#     old_value = cell.value
#     old_color = cell.font.name
#
#     if old_value != value and old_value is not None:
#         cell.value = value
#         cell.font = Font(
#
#         )


def multiMissionMain(pattern_data: patternInfo, row: Tuple[Cell, ...]):
    global MAX_COL_USED, REQUEST_DATA, CELL_WRITE_LOCK, GREEN, YELLOW

    cleaned_response = searchMail(search_params=pattern_data, mail_box="ALL")

    print([mail.subject for mail in cleaned_response])

    if not cleaned_response:
        print("未找到:", row[1].value)
        return None

    # 从邮件中提取最新版本信息
    newest_mail = cleaned_response[0] if cleaned_response else None
    newest_matched_data = MAIN_RE.match(newest_mail.subject).groupdict() if newest_mail else None
    print(newest_matched_data)

    # 写入数据缓存
    write_data: dict[str, Optional[str]] = {}
    workflow_data: List[dict] = []

    # 版本号及上传时间
    write_data['ver'] = newest_matched_data['ver'] if newest_matched_data else ''
    write_data['sentDate'] = newest_mail.SentDate.date().isoformat() if newest_mail else ''

    # 查询工作流
    if not newest_matched_data['ver'].isdigit() and newest_matched_data['wf']:
        # 工作流编号
        write_data['wf'] = newest_matched_data['wf'] if newest_matched_data else ''

        workflows_data = searchWorkflow(workflow_num=newest_matched_data['wf'])
        for workflow in workflows_data.workflows:
            # print(
            #     f"Workflow ID: {workflow.workflow_id}, Step Status: {workflow.step_status}, Step Name: {workflow.step_name}, "
            #     f"Step Out Come: {workflow.step_outcome}, Assignee: {workflow.assignees[0].name} "
            #     # f"Organization Name: {workflow.assignees[0].organization_name}"
            # )
            if workflow.step_name == "最终" and workflow.step_outcome != "正等待处理":
                # 审核完成，写入最终审核状态
                write_data['final_status'] = f"code {workflow.step_outcome.split('-')[0]}" if workflow.step_status != "已终止" else "工作流已终止"
            if workflow.step_outcome == "正等待处理":
                # 正在处理的工作流，写入处理人和状态
                workflow_data.append({
                    "org": workflow.assignees[0].organization_name.split(" ")[0],
                    "name": workflow.assignees[0].name.split(" ")[-1],
                    "status": workflow.step_status
                })

    # 完成请求，写入 REQUEST_DATA
    response_data = {
        "unit": matched_data['unit'],
        "discipline": matched_data['discipline'],
        "drawing": matched_data['drawing'],
        "wf": newest_matched_data['wf'] if newest_matched_data else None,
        "ver": newest_matched_data['ver'] if newest_matched_data and newest_matched_data['ver'] else "*",
        "title": newest_matched_data['title'] if newest_matched_data and newest_matched_data['title'] else None,
    }

    with CELL_WRITE_LOCK:
        # 清理审批结果、工作流编号、审批进度信息
        for a in row[6:]:
            a.value = None

        # 写入单元格
        row[4].value = write_data.get('ver', '')
        row[5].value = write_data.get('sentDate', '')
        row[6].value = write_data.get('final_status', '')
        row[7].value = write_data.get('wf', '')

        # 写入工作流数据
        base_col = 8    # 审核人起始列号

        for wf in workflow_data:
            row[base_col].value = wf['org']
            row[base_col + 1].value = wf['name']
            row[base_col + 2].value = wf['status']
            base_col += 3

        # 写入样式
        for b in row[:8]:
            if write_data.get('ver').isdigit():
                b.fill = PatternFill("solid", fgColor=GREEN)
            elif not write_data.get('ver').isdigit() and write_data.get('final_status') in ["code 3", "code 4"]:
                b.fill = PatternFill("solid", fgColor=YELLOW)
            elif not write_data.get('ver').isdigit() and not write_data.get('final_status') and not write_data.get('wf'):
                b.fill = PatternFill("solid", fgColor=BLUE)
            else:
                b.fill = PatternFill()  # no fill

        # 更新表MAX_COL_USED
        if base_col > MAX_COL_USED:
            MAX_COL_USED = base_col

        # 写入全局变量
        REQUEST_DATA[sheet.title].results.append(patternInfo(**response_data))
        REQUEST_DATA[sheet.title].total += 1
        REQUEST_DATA[sheet.title].unfinished += 1 if newest_matched_data and not newest_matched_data['ver'].isdigit() else 0

    return None


if __name__ == '__main__':
    # check input/export path
    if not os.path.isfile(XLSX_PATH):
        raise FileNotFoundError(f"Input file '{XLSX_PATH}' not found.")
    if os.path.isfile(EXPORT_PATH) and not EXPORT_PATH == XLSX_PATH:
        print(f"Warning: Output file '{EXPORT_PATH}' already exists and will be overwritten.")
        os.remove(EXPORT_PATH)

    # ensure access token is valid
    requestToken()

    # open and process xlsx
    wb = openpyxl.load_workbook(XLSX_PATH)
    for sheet in wb.worksheets:
        if sheet.title in ["汇总"]:  # 跳过汇总表
            continue

        # 初始化REQUEST_DATA
        REQUEST_DATA[sheet.title] = searchResult(sheet_name=sheet.title)

        # 构造线程池
        pool = ThreadPoolExecutor(max_workers=50)
        all_tasks = []

        # 遍历行，跳过表头
        for row in sheet.iter_rows(min_row=2, max_col=50):
            if row[1].value is None:
                continue
            m = MAIN_RE.match(clean_str(row[1].value))
            if not m:
                print("无法匹配:", row[1].value)
                continue

            matched_data = m.groupdict()
            # multiMissionMain(pattern_data=patternInfo(unit=m["unit"], discipline=m["discipline"], drawing=m["drawing"]), row=row)
            all_tasks.append(pool.submit(multiMissionMain, patternInfo(
                unit=matched_data['unit'], discipline=matched_data['discipline'], drawing=matched_data['drawing']), row))

        wait(all_tasks, return_when=ALL_COMPLETED)
        pool.shutdown()

        # 计算使用过的单元格最大数值，添加边框
        thin_side = Side(border_style="thin", color="000000")
        for _row in sheet.iter_rows(min_row=1, max_col=50):
            for _cell in _row[:MAX_COL_USED]:
                if type(_cell) is not MergedCell:
                    _cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
            for _cell in _row[MAX_COL_USED:]:
                if type(_cell) is not MergedCell:
                    _cell.border = Border()
                    _cell.value = None
                    _cell.fill = PatternFill()

        # 动态调整表头
        headers_group = ["待审批单位", "审批人", "审批状态"]

        # 从第 8 列开始，写入直到 MAX_COL_USED
        for col in range(9, MAX_COL_USED + 1, 3):
            for offset, title in enumerate(headers_group):
                sheet.cell(row=1, column=col + offset, value=title)

        # 清理超出 MAX_COL_USED 的表头
        for col in range(MAX_COL_USED + 1, 51):
            sheet.cell(row=1, column=col, value=None)

        wb.save(EXPORT_PATH)

        print(rf"Sheet '{sheet.title}' processed, total: {REQUEST_DATA[sheet.title].total}, unfinished: {REQUEST_DATA[sheet.title].unfinished}.")

    # 读取各子表审核进度，写入汇总sheet
    summary_sheet = wb["汇总"]
    for row in summary_sheet.iter_rows(min_row=2, max_col=6):
        # 写入汇总表
        if row[1].value is None:
            continue
        sheet_name = clean_str(row[1].value)
        if sheet_name not in REQUEST_DATA:
            print(f"Warning: Sheet '{sheet_name}' not found in processed data.")
            continue
        row[2].value = REQUEST_DATA[sheet_name].total
        row[4].value = REQUEST_DATA[sheet_name].unfinished

        # sheet tab 添加颜色
        if REQUEST_DATA[sheet_name].unfinished == 0:
            wb[sheet_name].sheet_properties.tabColor = GREEN

    wb.save(EXPORT_PATH)
    wb.close()
